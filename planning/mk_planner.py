# -*- coding: utf-8 -*-
"""Builds Pre_Day_1_Portal_Alignment_Planner.xlsx"""
import sys
sys.path.insert(0, '/tmp/claude-0/-home-user-ONBD-portal-prototype-/b3f9c803-e234-5ca7-a4f0-7c2389865b4f/scratchpad')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from mk_planner_part1 import PORTAL_MAP
from mk_planner_part2 import PORTAL_MAP_HM
from mk_planner_part3 import ALIGNMENT, DOCUMENTS

# ---- Equinix brand palette (from the Brand Center pack) ----
CHARCOAL = "2F3541"; CARBON = "5A657B"; CLOUD = "E1E3E7"; SILVER = "F2F3F4"
RED = "E91C24"; LT_BLUE = "CCE3FF"; DK_BLUE = "00408C"; LT_GREEN = "DFFBE5"
DK_GREEN = "2A8346"; LT_YELLOW = "FFF1CC"; LT_RED = "FFEBEE"; DK_RED = "AD050C"
LT_VIOLET = "E9DBFF"; DK_VIOLET = "411980"; LT_AQUA = "C7FDFF"; DK_AQUA = "00737A"

FONT = "Arial"
thin = Side(style="thin", color=CLOUD)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, ncols, fill=CHARCOAL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30

def title_block(ws, title, subtitle, ncols):
    ws.cell(row=1, column=1, value=title).font = Font(name=FONT, bold=True, size=15, color=CHARCOAL)
    ws.cell(row=2, column=1, value=subtitle).font = Font(name=FONT, size=9.5, color=CARBON)
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(ncols, 8))
    ws.row_dimensions[2].height = 30

def write_rows(ws, rows, start, widths, colour_col=None, colour_map=None):
    for ri, row in enumerate(rows, start=start):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name=FONT, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
        if ri % 2 == 0:
            for ci in range(1, len(widths) + 1):
                ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor=SILVER)
        if colour_col and colour_map:
            v = str(row[colour_col - 1])
            for key, (fill, fg) in colour_map.items():
                if v.startswith(key):
                    c = ws.cell(row=ri, column=colour_col)
                    c.fill = PatternFill("solid", fgColor=fill)
                    c.font = Font(name=FONT, size=9, bold=True, color=fg)
                    break
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

wb = Workbook()

# =====================================================================
# 1 — READ ME
# =====================================================================
ws = wb.active; ws.title = "Read Me"
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 118
ws.cell(row=1, column=1, value="Pre-Day 1 Portal — Alignment Planner").font = Font(name=FONT, bold=True, size=16, color=CHARCOAL)
ws.cell(row=2, column=1, value="Maps the built prototypes against every source we hold, and states what is aligned, what closed, and what is still open.").font = Font(name=FONT, size=10, color=CARBON)

readme = [
 ("", ""),
 ("WHAT THIS IS", ""),
 ("Purpose", "One place to see what the New Hire and Hiring Manager portal mockups actually contain, field by field, and how each part stands against the specifications, the future-state workflow, the commonality analysis and what Workday serves today. Built so the next conversation is about decisions, not about finding things."),
 ("Counts", "Every COUNT row in this workbook is a live formula, so the numbers stay right if you add or re-verdict rows. They calculate when you open the file."),
 ("What it is not", "Not a build spec and not an approved design. The mockups are a gap-finding tool; this planner is the audit of what they found."),
 ("Scope", "Pre-Day 1 only, two personas. People Experience is named throughout as an audience but has no screens and no UI spec available here — logged as G-19."),
 ("", ""),
 ("HOW TO READ THE TABS", ""),
 ("Portal Map", "Every screen, tab, section and field in the v3 mockups, both sides. Use the Status column to see where each element stands, and the Connects to column to see which elements cross between the two portals."),
 ("Alignment & Gaps", "The comparison. One row per finding, with what the mockup does, what the source says, a verdict and the action. Sort by Priority to get the working list."),
 ("Pre-Day 1 Documents", "What we ask new hires to read and sign today, from the live Workday business process, mapped against the intended future treatment and against what the mockup shows. This is where the largest single gap sits."),
 ("Consideration (next)", "Not built yet. Reserved for knitting in the handbook guide and the Inside Equinix content — see G-09, which now has a sourced answer to build from."),
 ("", ""),
 ("VERDICTS USED", ""),
 ("ALIGNED", "The mockup and the source agree. No action beyond monitoring."),
 ("GAP CLOSED", "A question the mockup had marked as an assumption now has a sourced answer. The mockup may still need to change to match it."),
 ("GAP OPEN", "Still unresolved, and someone has to decide."),
 ("CONFLICT", "Two or more sources disagree with each other. These are the most expensive to leave alone, because both readings are currently being built towards."),
 ("SUPERSEDED", "The mockup's position is out of date and should be replaced."),
 ("NEW REQUIREMENT", "Arrived in a source after the mockup was built and is not represented yet."),
 ("", ""),
 ("PRIORITY", ""),
 ("1 — Blocks build", "A developer cannot proceed past this. Five rows."),
 ("2 — Build now", "Answerable now, and the mockup should change once answered."),
 ("3 — Content", "The mechanism is right; someone needs to supply real text or values."),
 ("4 — Monitor", "Aligned today; worth re-checking if a source moves."),
 ("", ""),
 ("SOURCES USED", ""),
 ("Built artifacts", "New Hire portal mockup (v3) and Hiring Manager portal mockup (v3), sharing one state so cross-portal handoffs are real. 74-entry assumption register: 47 new hire, 18 manager, 9 connection."),
 ("Specifications", "Pre_Day_1_Task_UI_Spec.xlsx (New Hire) · HM_Pre_Day_1_Portal_UI_Spec.xlsx (Hiring Manager). No PEX equivalent available."),
 ("Requirements", "Equinix_Onboarding_Reimagined_PRD v1.3 and v1.4. v1.4 is the operative version and adds the corporate card flow, the Data Protection Policy language rule, the Payroll/ADP universal form, the proxy requirement and the Inside Equinix carousel."),
 ("Future state", "Onboarding_Portal_Future_State_Workflow_Template_v3 (Clean Day1) — Pre-Day 1, All Tasks, Key Principles, Decision Log."),
 ("Commonality", "Onboarding_Artifacts_Commonality_Workbook — Categorization Guide, Design Logic Tracker, Handbook Analysis, Artifact Inventory."),
 ("Current state", "CR_PC_Onboarding_Workflow_Documents — the live Workday business process document list. 146 rows, 416 document instances, 357 unique titles."),
 ("Supporting", "Global DPN pack (20 languages) · Credit Card Employee User Agreement · Hiring Manager Checklist (One Legal, 2019) · Buddy Program deck · Equipment provisioning deck · Onboarding Workday capability deck · Global New Hire Swag."),
 ("", ""),
 ("THE FIVE THINGS THAT MATTER MOST", ""),
 ("1", "The compliance pack is far larger than the mockup shows. Six global-core documents on screen; 416 document instances in the live process, with a 28-document bundle for every US hire and ~57 for Poland. (G-01, G-24, G-25)"),
 ("2", "Document language is derived from country, not chosen by the employee. The mockup has this the wrong way round. (G-02)"),
 ("3", "The corporate card is now fully specified and should be built — manager yes/no pre-start, new hire e-signature on Day 2. (G-06)"),
 ("4", "Two things built as net-new — the suggested network and the welcome note — may already exist as native platform capability. Re-cost before building. (G-12, G-13)"),
 ("5", "Two required new hire tasks are missing from the mockup entirely: confirm start date, and launching the background check. (G-21, G-22)"),
]
r = 4
for k, v in readme:
    a = ws.cell(row=r, column=1, value=k)
    b = ws.cell(row=r, column=2, value=v)
    if k and not v:
        a.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        a.fill = PatternFill("solid", fgColor=CHARCOAL)
        b.fill = PatternFill("solid", fgColor=CHARCOAL)
    else:
        a.font = Font(name=FONT, bold=True, size=9, color=CHARCOAL)
        b.font = Font(name=FONT, size=9, color=CARBON)
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    r += 1
ws.freeze_panes = "A4"

# =====================================================================
# 2 — PORTAL MAP
# =====================================================================
ws = wb.create_sheet("Portal Map")
cols = ["Side","Screen ID","Screen","Route","Section / Tab","Field or element","Type","Required",
        "Prefill / source","Behaviour & conditional logic","Connects to (other portal)",
        "Assumption","Spec reference","Status"]
title_block(ws, "Portal Map — every screen, tab and field in the mockups",
            "Both portals as built. 'Connects to' names the element on the other side that the same underlying fact drives — those are the handoffs. "
            "'Status' is the position against the sources; the reasoning for anything not Aligned is on the Alignment & Gaps tab.", len(cols))
ws.append([]); ws.append(cols)
style_header(ws, 4, len(cols))
rows = list(PORTAL_MAP) + list(PORTAL_MAP_HM)
status_map = {
 "Aligned": (LT_GREEN, DK_GREEN), "Gap closed": (LT_AQUA, DK_AQUA),
 "Open": (LT_YELLOW, "8A5A00"), "CONFLICT": (LT_RED, DK_RED),
 "Mockup only": (CLOUD, CARBON), "Mockup device": (CLOUD, CARBON),
}
write_rows(ws, rows, 5, [13,10,26,15,17,34,15,11,20,44,20,11,20,20],
           colour_col=14, colour_map=status_map)
ws.freeze_panes = "C5"
ws.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{4+len(rows)}"

# summary counts as live formulas
s = 6 + len(rows)
ws.cell(row=s, column=1, value="COUNTS").font = Font(name=FONT, bold=True, size=10, color=CHARCOAL)
last = 4 + len(rows)
for i, (lbl, formula) in enumerate([
    ("Elements mapped", f'=COUNTA(F5:F{last})'),
    ("New hire side", f'=COUNTIF(A5:A{last},"New hire")'),
    ("Hiring manager side", f'=COUNTIF(A5:A{last},"Hiring manager")'),
    ("Elements with a cross-portal link", f'=COUNTIFS(K5:K{last},"<>—",K5:K{last},"<>")'),
    ("Aligned", f'=COUNTIF(N5:N{last},"Aligned")'),
    ("Gap closed", f'=COUNTIF(N5:N{last},"Gap closed*")'),
    ("Open", f'=COUNTIF(N5:N{last},"Open*")'),
    ("Conflict", f'=COUNTIF(N5:N{last},"CONFLICT*")'),
]):
    ws.cell(row=s+1+i, column=1, value=lbl).font = Font(name=FONT, size=9, color=CARBON)
    c = ws.cell(row=s+1+i, column=2, value=formula)
    c.font = Font(name=FONT, bold=True, size=9, color=DK_BLUE)

# =====================================================================
# 3 — ALIGNMENT & GAPS
# =====================================================================
ws = wb.create_sheet("Alignment & Gaps")
cols = ["Ref","Area","Screen(s)","What the mockup does today","Source of truth",
        "What the source says","Verdict","Action needed","Owner","Priority"]
title_block(ws, "Alignment & Gaps — mockups against the sources",
            "One row per finding. Sort by Priority for the working list. Six rows block build. Four are conflicts between sources, which are the expensive ones "
            "because both readings are currently being designed towards.", len(cols))
ws.append([]); ws.append(cols)
style_header(ws, 4, len(cols))
verdict_map = {
 "ALIGNED": (LT_GREEN, DK_GREEN), "GAP CLOSED": (LT_AQUA, DK_AQUA),
 "GAP OPEN": (LT_YELLOW, "8A5A00"), "CONFLICT": (LT_RED, DK_RED),
 "SUPERSEDED": (LT_VIOLET, DK_VIOLET), "NEW REQUIREMENT": (LT_BLUE, DK_BLUE),
}
write_rows(ws, ALIGNMENT, 5, [8,26,20,50,30,62,16,52,18,15],
           colour_col=7, colour_map=verdict_map)
# priority colouring
for ri in range(5, 5 + len(ALIGNMENT)):
    v = str(ws.cell(row=ri, column=10).value or "")
    c = ws.cell(row=ri, column=10)
    if v.startswith("1"): c.fill = PatternFill("solid", fgColor=LT_RED); c.font = Font(name=FONT, size=9, bold=True, color=DK_RED)
    elif v.startswith("2"): c.fill = PatternFill("solid", fgColor=LT_YELLOW); c.font = Font(name=FONT, size=9, bold=True, color="8A5A00")
    elif v.startswith("3"): c.fill = PatternFill("solid", fgColor=LT_BLUE); c.font = Font(name=FONT, size=9, bold=True, color=DK_BLUE)
    else: c.fill = PatternFill("solid", fgColor=SILVER); c.font = Font(name=FONT, size=9, color=CARBON)
ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{4+len(ALIGNMENT)}"
last = 4 + len(ALIGNMENT)
s = 6 + len(ALIGNMENT)
ws.cell(row=s, column=1, value="COUNTS").font = Font(name=FONT, bold=True, size=10, color=CHARCOAL)
for i, (lbl, formula) in enumerate([
    ("Findings", f'=COUNTA(A5:A{last})'),
    ("Blocks build (P1)", f'=COUNTIF(J5:J{last},"1*")'),
    ("Build now (P2)", f'=COUNTIF(J5:J{last},"2*")'),
    ("Conflicts between sources", f'=COUNTIF(G5:G{last},"CONFLICT")'),
    ("Answered by these documents", f'=COUNTIF(G5:G{last},"GAP CLOSED")+COUNTIF(G5:G{last},"SUPERSEDED")+COUNTIF(G5:G{last},"NEW REQUIREMENT")'),
    ("Still open", f'=COUNTIF(G5:G{last},"GAP OPEN")'),
    ("Aligned", f'=COUNTIF(G5:G{last},"ALIGNED")'),
]):
    ws.cell(row=s+1+i, column=1, value=lbl).font = Font(name=FONT, size=9, color=CARBON)
    c = ws.cell(row=s+1+i, column=2, value=formula)
    c.font = Font(name=FONT, bold=True, size=9, color=DK_BLUE)

# =====================================================================
# 4 — PRE-DAY 1 DOCUMENTS
# =====================================================================
ws = wb.create_sheet("Pre-Day 1 Documents")
cols = ["Document family","Document(s) served today","Who receives it","Instances in the Workday process",
        "Acknowledged today?","Intended future treatment","In the v3 mockup?","Gap / action"]
title_block(ws, "Pre-Day 1 Documents — what we ask every new hire to read and sign today",
            "Left of the line is the live Workday business process; right of it is the intended future treatment from the commonality analysis, and what the "
            "mockup currently shows. Totals across the whole process: 146 rows, 416 document instances, 357 unique titles. The six global-core documents in the "
            "mockup are the intended future core — not what a new hire receives today.", len(cols))
ws.append([]); ws.append(cols)
style_header(ws, 4, len(cols))
mock_map = {
 "Yes": (LT_GREEN, DK_GREEN), "Partly": (LT_YELLOW, "8A5A00"),
 "No": (LT_RED, DK_RED), "Shown": (LT_AQUA, DK_AQUA), "Not represented": (LT_RED, DK_RED),
 "Referenced": (LT_GREEN, DK_GREEN),
}
write_rows(ws, DOCUMENTS, 5, [26,54,30,16,16,52,38,58], colour_col=7, colour_map=mock_map)
ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{4+len(DOCUMENTS)}"

s = 6 + len(DOCUMENTS)
notes = [
 ("SCALE OF THE CURRENT STATE", ""),
 ("Document instances served", "416 across 146 business-process rows"),
 ("Unique document titles", "357"),
 ("Global core (all countries)", "6 — Code of Business Conduct, Company Factsheet, Acceptable Use Policy, Information Governance, Securities Trading, Whistleblower"),
 ("Largest country pack", "Poland, ~57 instances — ahead of USA ~36, Brazil ~28, Japan ~19"),
 ("US handbook bundle", "28 documents in one row: the handbook, 25 state addenda, the benefits booklet and GovDocs — served to every US hire regardless of work state"),
 ("US state-conditional rows", "5 exist (CA, NJ, TX, MA, NY), which proves conditionality is possible in the same process"),
 ("", ""),
 ("READING THIS TAB", ""),
 ("Why the counts matter", "The mockup shows a six-item list that takes about ten minutes. If the real pre-Day 1 pack is closer to the current state, the screen is a different design problem — and Poland, not the US, is the one to draw."),
 ("What is genuinely settled", "The global core six, the one-global-notice-plus-country-addendum pattern for the DPN and AUP, and that the handbook is reference rather than an acknowledgement."),
 ("What is not", "Which of the 357 titles survive, which move past Day 1, and which are retired. That decision is the single biggest open item in this planner."),
]
for i, (k, v) in enumerate(notes):
    a = ws.cell(row=s+i, column=1, value=k); b = ws.cell(row=s+i, column=2, value=v)
    if k and not v:
        a.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        a.fill = PatternFill("solid", fgColor=CHARCOAL)
        for cc in range(2, 5): ws.cell(row=s+i, column=cc).fill = PatternFill("solid", fgColor=CHARCOAL)
    else:
        a.font = Font(name=FONT, bold=True, size=9, color=CHARCOAL)
        b.font = Font(name=FONT, size=9, color=CARBON)
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=s+i, start_column=2, end_row=s+i, end_column=6)

# =====================================================================
# 5 — CONSIDERATION (placeholder, deliberately)
# =====================================================================
ws = wb.create_sheet("Consideration (next)")
ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 110
ws.cell(row=1, column=1, value="Consideration — handbook guide and Inside Equinix").font = Font(name=FONT, bold=True, size=15, color=CHARCOAL)
ws.cell(row=2, column=1, value="Deliberately empty. This is the next tab to build, once the alignment above has been reviewed.").font = Font(name=FONT, size=10, color=CARBON)
seed = [
 ("", ""),
 ("WHAT ALREADY HAS AN ANSWER", ""),
 ("Inside Equinix placement", "PRD v1.4 now specifies it: a carousel at the top of the portal rotating chapters 01-06, plus a right-hand bar link under Additional Reading. No tasks — access and the carousel only. The same content also appears in New Hire Orientation and in the manager's weekly task list. See G-09."),
 ("Company factsheet", "Single global object, English only, no acknowledgement — settled. The open question is whether it belongs pre-Day 1 at all or on Day 1."),
 ("Employee handbook", "Country-conditional reference, no acknowledgement, available Pre-Day 1 through Day 30. Key policies acknowledged separately as universal tasks."),
 ("Manager Companion Guide", "Referenced in PRD v1.4 as the content linked from the manager's weekly task list. Its relationship to the Manager Onboarding Guide is still unresolved (M-15)."),
 ("", ""),
 ("WHAT THIS TAB WILL NEED TO WORK OUT", ""),
 ("Sequencing", "Which of the six chapters a new hire is ready to absorb, and when. The PRD asks for a rotating carousel; the sequencing question is whether rotation is the right model or whether chapters should be released against the countdown."),
 ("Weight", "Culture content competes with compliance tasks for the same pre-Day 1 window. The portal currently opens with a task list; a carousel above it changes what the screen is for."),
 ("Duplication", "The same content appears in three places — the portal, orientation and the manager's list. Worth deciding whether that is reinforcement or repetition."),
 ("Handbook vs guide", "The country handbook is reference material with statutory content; Inside Equinix is culture. They are currently both 'reading material' in the mockup, which flattens a real difference."),
 ("Measurement", "Nothing is tracked, by design. Worth confirming that is still wanted, given the platform natively supports optional to-dos with no due date (A-36)."),
]
r = 4
for k, v in seed:
    a = ws.cell(row=r, column=1, value=k); b = ws.cell(row=r, column=2, value=v)
    if k and not v:
        a.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        a.fill = PatternFill("solid", fgColor=CHARCOAL); b.fill = PatternFill("solid", fgColor=CHARCOAL)
    else:
        a.font = Font(name=FONT, bold=True, size=9, color=CHARCOAL)
        b.font = Font(name=FONT, size=9, color=CARBON)
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    r += 1

out = "/home/user/ONBD-portal-prototype-/planning/Pre_Day_1_Portal_Alignment_Planner.xlsx"
import os
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("wrote", out)
print("portal map rows:", len(rows), "| alignment rows:", len(ALIGNMENT), "| document rows:", len(DOCUMENTS))
